import re, sys, json,datetime,openpyxl, platform
from datetime import datetime
from datetime import datetime,date
import time, random#, testlink
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException, WebDriverException
from selenium.webdriver.support import expected_conditions as EC
from random import choice
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import Color, Fill

from openpyxl.cell import Cell
import pathlib 
from pathlib import Path
import os
from sys import platform
import H_function
import select
from openpyxl import Workbook
# from simple_colors import *



chrome_options = webdriver.ChromeOptions()






#result=open(os.path.dirname(Path(__file__).absolute())+'\\result.txt','w')

#print(json_file)
class objects:
    now = datetime.now()
    year = now.strftime("%Y")
    month = now.strftime("%m")
    day = now.strftime("%d")
    time1 = now.strftime("%H:%M:%S")
    date_time = now.strftime("%Y/%m/%d, %H:%M:%S")
    date_id = date_time.replace("/", "").replace(", ", "").replace(":", "")[2:]
#XlsxName="TestCaseAllMenu"+ str(random.randint(0,10000))+".xlsx"
XlsxName="TestCaseAllMenu"+ str(objects.date_id) +".xlsx"
if platform == "linux" or platform == "linux2":
    local_path = "/home/oem/groupware-auto-test"
    json_file = local_path + "/H_groupware_auto.json"
    with open(json_file) as json_data_file:
        data = json.load(json_data_file)
    driver = webdriver.Chrome("/usr/bin/chromedriver")
    log_folder = "/Log/"
    log_testcase = "/Log/"
    xlsx_xpath =os.path.dirname(Path(__file__).absolute())+log_folder+XlsxName
    execution_log = local_path + log_folder + "execution_log_" + str(objects.date_id) + ".txt"
    # fail_log = execution_log.replace("execution_log_", "fail_log_")
    # error_log = execution_log.replace("execution_log_", "error_log_")
    testcase_log = local_path + log_testcase + "Hanh_TestcaseAllmenu_" + str(objects.date_id) + ".xlsx"
    
else:
    local_path = os.path.dirname(Path(__file__).absolute())
    json_file = local_path + "\\H_groupware_auto.json"
    with open(json_file) as json_file:
        data = json.load(json_file)
    driver = webdriver.Chrome(local_path + "\\chromedriver.exe")
    json_file = local_path + "\\H_groupware_auto.json"
    log_folder = "\\Log\\"
    log_testcase = "\\Log\\"
    xlsx_xpath =os.path.dirname(Path(__file__).absolute())+'\\'+XlsxName
    execution_log = local_path + log_folder + "execution_log_" + str(objects.date_id) + ".txt"
    # fail_log = execution_log.replace("execution_log_", "fail_log_")
    # error_log = execution_log.replace("execution_log_", "error_log_")
    testcase_log = local_path + log_testcase + "Hanh_TestcaseAllmenu_" + str(objects.date_id) + ".xlsx"

driver.maximize_window()
driver.implicitly_wait(10)

logs = [testcase_log, execution_log]#, error_log, fail_log]
for log in logs:
    if ".txt" in log:
        open(log, "x").close()
    else:
        wb = Workbook()
        myFill = PatternFill(start_color='adc5e7',
                   end_color='adc5e7',
                   fill_type='solid',)
        font = Font(name='Calibri',
                    size=11 ,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
        ws = wb.active

        ws.cell(row=1, column=1).value= "Menu"
        ws.cell(row=1, column=2).value = "Sub-Menu"
        ws.cell(row=1, column=3).value = "Test Case Name"
        ws.cell(row=1, column=4).value = "Status"
        ws.cell(row=1, column=5).value = "Description"
        ws.cell(row=1, column=6).value = "Date"
        ws.cell(row=1, column=7).value = "Tester"
        # color 
        ws.cell(row=1, column=1).fill = myFill
        ws.cell(row=1, column=2).fill = myFill
        ws.cell(row=1, column=3).fill = myFill
        ws.cell(row=1, column=4).fill = myFill
        ws.cell(row=1, column=5).fill = myFill
        ws.cell(row=1, column=6).fill = myFill
        ws.cell(row=1, column=7).fill = myFill
        # font
        ws.cell(row=1, column=1).font = Font(bold=True)
        ws.cell(row=1, column=2).font = Font(bold=True)
        ws.cell(row=1, column=3).font = Font(bold=True)
        ws.cell(row=1, column=4).font = Font(bold=True)
        ws.cell(row=1, column=5).font = Font(bold=True)
        ws.cell(row=1, column=6).font = Font(bold=True)
        ws.cell(row=1, column=7).font = Font(bold=True)

        wb.save(log)

def Logging(msg):
    print(msg)
    log_msg = open(execution_log, "a")
    log_msg.write(str(msg) + "\n")
    log_msg.close()

def TesCase_LogResult(menu, sub_menu, testcase, status, description, tester):
    #Logging(description)

    # if status == "Pass":
    #     print(objects.testcase_pass)
    # else:
    #     print(objects.testcase_fail)

    wb = openpyxl.load_workbook(testcase_log)
    current_sheet = wb.active
    start_row = len(list(current_sheet.rows)) + 1

    current_sheet.cell(row=start_row, column=1).value = menu
    current_sheet.cell(row=start_row, column=2).value = sub_menu
    current_sheet.cell(row=start_row, column=3).value = testcase
    current_sheet.cell(row=start_row, column=4).value = status
    current_sheet.cell(row=start_row, column=5).value = description
    current_sheet.cell(row=start_row, column=6).value = objects.date_time
    current_sheet.cell(row=start_row, column=7).value = tester

    wb.save(testcase_log)


# # create log file of fail test case
# open(execution_log, "x").close()

# # create log file of fail test case
# open(fail_log, "x").close()

# # create log file of fail test case
# open(error_log, "x").close()
'''
def Logging(msg):
    print(msg)
    log_msg = open(execution_log, "a")
    log_msg.write(str(msg) + "\n")
    log_msg.close()
'''


'''TESTLINK_API_PYTHON_SERVER_URL= 'http://qa1.hanbiro.net/testlink/lib/api/xmlrpc/v1/xmlrpc.php'
TESTLINK_API_PYTHON_DEVKEY= 'e80ce28cbff80bd624cc2679c915596d'

tls = testlink.TestLinkHelper(TESTLINK_API_PYTHON_SERVER_URL, TESTLINK_API_PYTHON_DEVKEY).connect(testlink.TestlinkAPIClient)

def TestlinkResult_Pass(external_id):
    tls.reportTCResult(testcaseexternalid=external_id, testplanid=7377, buildname="v3.8.45", status='p', notes='Test Case [' + external_id + '] passed')

def TestlinkResult_Fail(external_id):
    tls.reportTCResult(testcaseexternalid=external_id, testplanid=7377, buildname="v3.8.45", status='f', notes='Test Case [' + external_id + '] failed')'''

def access_hr(domain, id, pw):
    driver.get(domain + "/nhr/login")
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.NAME, "gw_id")))
    userID = driver.find_element_by_name("gw_id")
    # userID.send_keys(data["user_name"])
    userID.send_keys(id)
    print("- Input user ID")
    password = driver.find_element_by_name("gw_pass")
    # password.send_keys(data["user_password"])
    password.send_keys(pw)
    print("- Input user password")
    driver.find_element_by_xpath(data["TIMECARD"]["sign_in"]).click()
    print("- Click button Sign in")
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["TIMECARD"]["notify"][0])))
    print("=> Log in successfully")
    TesCase_LogResult(**data["testcase_result"]["HR-Timecard"]["checkin"]["pass"])
    time.sleep(2)
    driver.refresh()
    try:
        hr_project = driver.find_element_by_xpath("//h5[contains(.,'HR project')] ")
        if hr_project.is_displayed():
            driver.find_element_by_xpath("//h5[contains(.,'HR project')] ").click()
            time.sleep(3)
    except:
        Logging(" ")

    #driver.find_element_by_xpath("//*[@id='app']//a[contains(@href,'/nhr/hr')]").click()
    #WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["TIMECARD"]["body"])))
    time.sleep(5)
